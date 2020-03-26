#!/usr/bin/env python3

from jinja2 import Environment, FileSystemLoader, select_autoescape

import random
import string

header_height = 100
item_height = 50
item_width = 180
padding_bottom = 10
padding_left = 20
padding_right = 20
group_padding_left = 20
group_padding_right = 20
group_padding_bottom = 20


def read_team_data(data):
    td = {}
    header = None
    for row in data:
        # Get a list of all columns in each row
        cols = []
        for col in row:
            cols.append(col.value)
        
        if header is None:
            header = cols
            continue

        td[cols[0]] = {
            "layer": cols[2],
            "display_name": cols[1].replace("&", "&amp;"),
            "style": cols[3],
            "sort_order": cols[4],
        }
    
    return td


def read_configuration_data(wb):
    ws = wb["Configuration"]
    for tbl in ws._tables:
        if tbl.name == "Configuration":
            data = ws[tbl.ref]
            break

    cfg = {}
    header = None
    for row in data:
        # Get a list of all columns in each row
        cols = []
        for col in row:
            cols.append(col.value)
        
        if header is None:
            header = cols
        else:
            cfg[cols[0]] = cols[1]

    return cfg


def read_workgroup_data(data):
    wgs = {}
    header = None
    for row in data:
        # Get a list of all columns in each row
        cols = []
        for col in row:
            cols.append(col.value)
        
        if header is None:
            header = cols
            continue

        # skip empty line
        if cols[0] == "" or cols[0] is None:
            continue

        col3 = cols[3]
        if col3 is None:
            col3 = ""

        wgs[cols[0]] = {
            "type": cols[1],
            "team": cols[2],
            "members": col3.split(","),
        }
    return wgs


def read_module_data(data):
    mdls = []
    header = None
    for row in data:
        # Get a list of all columns in each row
        cols = []
        for col in row:
            cols.append(col.value)
        
        if header is None:
            header = cols
        else:
            mdl = {
                "team": cols[0],
                "group": cols[1],
                "display_name": cols[2],
                "wg_type": cols[3],
                "status": convert_status(cols[4]),
                "sub_group": cols[5],
                "type": "module",
            }
            mdls.append(mdl)
    return mdls


def convert_status(status):
    if status == "Not Started":
        return "StatusStyle0"
    if status == "In Progress":
        return "StatusStyle1"
    if status == "Completed":
        return "StatusStyle2"
    
    return "StatusStyleU"


def read_data(file_path):
    # Reading an excel file using Python 
    import openpyxl

    # To open Workbook 
    wb = openpyxl.load_workbook(file_path)

    cfg = read_configuration_data(wb)

    ws = wb["Teams"]
    
    # For row 0 and column 0 
    for tbl in ws._tables:
        # print(tbl.name)
        # Grab the 'data' from the table
        data = ws[tbl.ref]
        if tbl.name == "Teams":
            teams = read_team_data(data)

        if tbl.name == "WorkGroups":
            wgs = read_workgroup_data(data)
            # print(wgs)

        if tbl.name == "Modules":
            mdls = read_module_data(data)
            # print(mdls)

    # normalize data
    for mdl in mdls:
        mdl["wg_stype"] = "WGStyle%d" % (wgs[mdl["wg_type"]]["type"])


    return mdls, teams, wgs, cfg


def randomString(stringLength=10):
    """Generate a random string of fixed length """
    letters = string.ascii_lowercase
    return ''.join(random.choice(letters) for i in range(stringLength))


def layout_items(items, start_x, start_y):
    x = start_x
    y = start_y
    column = 0
    row = 0
    width = 0
    height = 0
    overflow_w = 0
    for item in items:
        item["x"] = x
        item["y"] = y
        column = (column + 1) % 3
        if column == 0:
            row = row + 1
            y += item_height + padding_bottom
            x = start_x
            overflow_w = item_width
        else:
            x += item_width + padding_left

        width = max(width, x - start_x)
        height = max(height, y - start_y)
    
    if column == 0:
        # full row
        height -= item_height + padding_bottom
    if overflow_w == 0 and column != 0:
        overflow_w = -padding_left

    width += overflow_w
    height += item_height

    return width, height


def relayout_items(items, start_x, end_y):
    y = end_y
    column = 0
    min_x = 10000000
    for item in items:
        min_x = min(min_x, item["x"])

    offset_x = min_x - start_x
    for item in items:
        item["x"] = item["x"] - offset_x
        item["y"] = y - item_height
        column = (column + 1) % 3
        if column == 0:
            y -= item_height + padding_bottom


def distribute_horizontal(items, start_x, padding):
    # sort items by sort order
    items.sort(key=lambda x: x['order'])
    for item in items:
        item["x"] = start_x
        start_x += item["w"] + padding


def move_workgroup_to(wg, start_x, start_y):
    wg["team0"]["x"] = start_x
    wg["team0"]["y"] = start_y

    offset_y = 40
    for i in range(1, 5):
        wgi = wg["team%d" % i]
        wgi["x"] = start_x
        wgi["y"] = start_y + offset_y
        offset_y += wgi["h"]


def layout_workgroup(wgs):
    height_lut = [30, 40, 50, 60, 70, 90, 110, 120, 140, 150, 160, 170]
    by_team = {}
    for wg in wgs.values():
        if wg["team"] not in by_team:
            by_team[wg["team"]] = { 
                "layout": {},
                "team0": {},
                "team1": {},
                "team2": {},
                "team3": {},
                "team4": {},
            }
        team = by_team[wg["team"]]
        team["team%d" % wg["type"]] = wg

    # create layout for each team
    for team in by_team.values():
        for i in range(5):
            ti = team["team%d" % i]
            ti["w"] = 110
            ti["h"] = height_lut[len(ti.get("members", []))]    # minimum height for empty
            ti["style"] = "WGStyle%d" % i
            ti["type"] = "wg"
            ti["id"] = randomString()
            ti["display_name"] = "&lt;br/&gt;".join(ti.get("members", ""))
        
        # recaliberate team 0 height
        h = 0
        for i in range(1, 5):
            ti = "team%d" % i
            h += team[ti]["h"]
        team["team0"]["h"] = h
        team["team0"]["display_name"] = "Leader: %s" % team["team0"]["display_name"]

        # start_x = 100
        # start_y = 200
        # move_workgroup_to(team, start_x, start_y)
        # print(team)

    return by_team


def create_layout(items, wgs_byteam, teams):
    root = {}
    layers = {}
    for item in items:
        team_code = item["team"]
        team_info = teams[team_code]
        if team_code not in root:
            root[team_code] = {
                "id": randomString(),
                "display_name": team_info["display_name"],
                "style": team_info["style"],
                "code": team_code,
                "type": "team",
                "order": team_info["sort_order"],
                "groups": {},
            }
        team = root[team_code]

        layer = team_info["layer"]
        if layer not in layers:
            layers[layer] = []

        r_layer = layers[layer]
        if team not in r_layer:
            r_layer.append(team)

        group_code = item["group"]
        if group_code not in team["groups"]:
            team["groups"][group_code] = {
                "id": randomString(),
                "display_name": group_code,
                "type":"group",
                "items":[],
            }
        group = team["groups"][group_code]
        item["id"] = randomString()
        group["items"].append(item)

    start_x = 0
    start_y = 300

    wg_padding = 180

    # first pass, layout all items inside group
    for team in root.values():
        team["x"] = start_x
        team["y"] = start_y
        max_w = 0
        max_h = 0
        for group in team["groups"].values():
            start_x = start_x + group_padding_left
            g_y = start_y + header_height
            w, h = layout_items(group["items"], start_x + padding_left, g_y + header_height)
            group["x"] = start_x
            group["y"] = g_y
            g_w =  w + group_padding_left + group_padding_right
            g_h = h + header_height + group_padding_bottom
            group["w"] = g_w
            group["h"] = g_h
            start_x = start_x + g_w
            max_w = max(max_w, start_x - team["x"])
            max_h = max(max_h, g_h)
        team["w"] = max_w + group_padding_right
        team["h"] = max_h + header_height + group_padding_bottom

        start_x = start_x + padding_left + padding_right # next team

    # second pass, rebalance teams, make sure all teams have equal height
    start_y = 0
    for l in sorted(layers.keys(), reverse = True):
        start_x = 120 + 80
        distribute_horizontal(layers[l], start_x, group_padding_right + wg_padding)

        max_h = 0
        for team in layers[l]:
            max_h = max(max_h, team["h"])

        for team in layers[l]:
            start_y = max(team["y"], start_y)
            team["y"] = start_y
            team["h"] = max_h
    
        start_y = start_y + max_h + group_padding_bottom * 3

    # max_h = 0
    # for team in root.values():
    #     max_h = max(max_h, team["h"])
    # for team in root.values():
    #     team["h"] = max_h

    # third pass, rebalance groups, make sure all groups have equal height
    for team in root.values():
        start_x = team["x"] + group_padding_left
        max_h = 0
        for group in team["groups"].values():
            max_h = max(max_h, group["h"])

        for group in team["groups"].values():
            h = team["h"] - header_height - group_padding_bottom
            start_y = team["y"] + header_height
            start_y = team["y"] + team["h"] - group_padding_bottom - max_h
            group["x"] = start_x
            group["y"] = start_y
            group["h"] = max_h

            relayout_items(group["items"], start_x + padding_left, start_y + max_h - group_padding_bottom)
            w = group["w"]
            start_x = start_x + w + group_padding_right

        move_workgroup_to(wgs_byteam[team["code"]], team["x"] - 120, team["y"])

    # flat out
    all_items = []
    for team in root.values():
        all_items.append(team)
        for group in team["groups"].values():
            all_items.append(group)
            all_items = all_items + group["items"]

    for wg in wgs_byteam.values():
        for i in range(5):
            all_items.append(wg["team%d" % i])

    # print(all_items)
    return all_items


def main():
    import argparse

    parser = argparse.ArgumentParser(description='Render team structure to drawio file format.')
    parser.add_argument('team_path', metavar='path', type=str,
                        help='path to xlsx file that contains team structure')

    args = parser.parse_args()

    env = Environment(
        loader=FileSystemLoader('./templates'),
        autoescape=select_autoescape(['html', 'xml']),
        trim_blocks = True,
        lstrip_blocks = True,
        line_statement_prefix='#',
    )

    template = env.get_template('corepayment.tmpl')

    mdls, teams, wgs, cfg = read_data(args.team_path)

    wgs_byteam = layout_workgroup(wgs)
    all = create_layout(mdls, wgs_byteam, teams)

    print(template.render(items=all, configs=cfg))


if __name__ == "__main__":
    main()
