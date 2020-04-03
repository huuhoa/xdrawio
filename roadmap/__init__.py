import xutils
import datetime

__version_info__ = (0,1,0)
__version__ = '.'.join([str(__value) for __value in __version_info__])
__copyright__ = '2020, NGUYEN Huu Hoa'
__license__ = 'MIT'


def read_column_data(wb):
    ws = wb["Banks"]
    for tbl in ws._tables:
        if tbl.name == "Stages":
            data = ws[tbl.ref]
            break

    td = []
    header = None
    for row in data:
        # Get a list of all columns in each row
        cols = []
        for col in row:
            cols.append(col.value)

        if header is None:
            header = cols
            continue

        td.append({header[i]:cols[i] for i in range(len(cols))})
    
    td.sort(key=lambda x: x["Sort Order"], reverse=False)
    return td


class Data(object):
    def __init__(self):
        super().__init__()
        self.column_stage = {}
        self.roadmap = {}
        self.styles = {}
        self.configurations = {}


def read_roadmap_data(wb):
    ws = wb["Roadmap"]
    for tbl in ws._tables:
        if tbl.name == "Roadmap":
            data = ws[tbl.ref]
            break

    d = {}
    header = None
    for row in data:
        # Get a list of all columns in each row
        cols = []
        for col in row:
            cols.append(col.value)

        if header is None:
            header = cols
            continue

        info = {header[i]:cols[i] for i in range(len(cols))}
        item = {}
        name = info["Feature Name"]
        item["id"] = xutils.randomString()
        item["display_name"] = info["Feature Name"]
        item["info"] = info
        d[name] = item

    return d


def read_data(file_path):
    # Reading an excel file using Python 
    import openpyxl

    d = Data()

    # To open Workbook 
    wb = openpyxl.load_workbook(file_path)

    d.configurations = xutils.read_configuration_data(wb)
    d.column_stage = read_column_data(wb)
    d.roadmap = read_roadmap_data(wb)

    return d

def convert_bank_status(status, configs):
    if status is None:
        return "NA"
    return configs.get(status, "Unknown")

def groupby_component(data):
    groups = {
        d["info"]["Component"]: {
                "name": d["info"]["Component"].replace(" ", ""),
                "display_name": d["info"]["Component"].upper(),
                "type": "component",
                "layers": {},
                "items": [],
            } for d in data.roadmap.values()
    }
    for d in data.roadmap.values():
        g = groups[d['info']["Component"]]
        items = g["items"]
        d["style"] = data.configurations["Priority_%s" % d['info']['Priority']]
        items.append(d)

    for g in groups.values():
        items = g["items"]
        layers = {
            d['info']['Layer']: [] for d in items
        }
        for d in items:
            l = layers[d['info']['Layer']]
            l.append(d)
        g["layers"] = layers

    return groups


def layout_group(groups, start_y, item_height, data):
    for g in groups.values():
        gname = g["name"]
        if gname is None:
            gname = "Not Yet"

        g["id"] = "group-%s" % g["name"]
        g["h"] = len(g["layers"]) * item_height + item_height - 10
        g["y"] = start_y
        g["style"] = data.configurations["Component_%s" % gname]

        start_y += g["h"] + 10


def compute_width_by_date(start_date):
    days_per_month = {
        1: 31,
        2: 28,
        3: 31,
        4: 30,
        5: 31,
        6: 30,
        7: 31,
        8: 31,
        9: 30,
        10: 31,
        11: 30,
        12: 31
    }
    if start_date.day == 1:
        return (start_date.month - 1) * 120

    offset_date = start_date.day - days_per_month[start_date.month]
    if offset_date == 0:
        return start_date.month * 120

    return (start_date.month - 1) * 120 + (start_date.day/days_per_month[start_date.month]) * 120


def date_is_nearby(d1, d2):
    delta = d1 - d2
    # print(d1, d2, delta.days)
    return delta.days >= -3 and delta.days <= 3


def flatten_roadmapitems(all_items, items, parentid, start_y):
    items.sort(key=lambda x: x['info']['End Date'], reverse=True)
    prev_item = {
        'Start Date': datetime.datetime(2000, 1, 1)
    }
    for b in items:
        start_date = b['info']['Start Date']
        end_date = b['info']['End Date']
        start = compute_width_by_date(start_date)
        end = compute_width_by_date(end_date)
        if date_is_nearby(end_date, prev_item['Start Date']):
            last_item = all_items[-1]
            last_item['start'] = last_item['start'] + 7

        prev_item = b['info']

        all_items.append({
            "id": b["id"],
            "y": start_y,
            "start": start + 80,
            "end": end + 80,
            "type": "item",
            "display_name": b["info"]["Feature Name"],
            "style": b['style'],
            "parentid": parentid,
        })


def flatten_data(data):
    all_items = []

    start_x = 0
    start_y = 200
    counter = 0
    stages = {}

    item_padding = 70
    item_height = 50
    groups = groupby_component(data)
    layout_group(groups, start_y, item_height, data)

    total_height = 0
    for g in groups.values():
        total_height += g["h"] + 10

    x = start_x

    for g in groups.values():
        all_items.append(g)
        inner_y = 50
        parentid = g["id"]
        for l in sorted(g["layers"].keys()):
            layer = g["layers"][l]
            flatten_roadmapitems(all_items, layer, parentid, inner_y)
            inner_y += item_height

    return all_items
