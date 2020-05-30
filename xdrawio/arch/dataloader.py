import xdrawio.xutils


def read_table_from_wb(wb, sheet_name, table_name):
    ws = wb[sheet_name]
    for tbl in ws._tables:
        if tbl.name == table_name:
            return ws[tbl.ref]

    return None


def transform_table_to_frame(table):
    ''' transform input (array of array) to output (array of dict)'''
    d = []
    header = None
    for row in table:
        # Get a list of all columns in each row
        cols = []
        for col in row:
            cols.append(col.value)

        if header is None:
            header = cols
            continue

        item = {header[i]: cols[i] for i in range(len(cols))}
        d.append(item)
    return d


def transform_table_to_dict(table, key_func):
    ''' transform input (array of array) to output (dict of dict)'''
    d = {}
    header = None
    for row in table:
        # Get a list of all columns in each row
        cols = []
        for col in row:
            cols.append(col.value)

        if header is None:
            header = cols
            continue

        item = {header[i]: cols[i] for i in range(len(cols))}
        key = key_func(item)
        d[key] = item
    return d


def read_domain_data(data):
    td = transform_table_to_dict(data, lambda item: item['Code'])
    for k, v in td.items():
        v['Domain Name'] = xdrawio.xutils.encode_name(v['Domain Name'])
    return td


def read_group_data(data):
    d = transform_table_to_dict(data, lambda item: item['Group Code'])
    for k, v in d.items():
        v['Group Name'] = xdrawio.xutils.encode_name(v['Group Name'])

    return d


class Data(object):
    def __init__(self):
        super().__init__()
        self.domains = {}
        self.groups = {}
        self.workgroups = {}
        self.configurations = {}
        self.layoutspec = {}


def read_layout_specification(data):
    d = {}
    for row in data:
        # Get a list of all columns in each row
        cols = []
        for col in row:
            cols.append(col.value)

        d[cols[0]] = cols[1]

    return d


def read_workgroup_data(data):
    wgs = {}
    header = None
    for row in data:
        # Get a list of all columns in each row
        cols = []
        for col in row:
            cols.append(col.value)

        if header is None:
            header = cols
            continue

        # skip empty line
        if cols[0] == "" or cols[0] is None:
            continue

        col3 = cols[3]
        if col3 is None:
            col3 = ""

        wgs[cols[0]] = {
            "type": cols[1],
            "team": cols[2],
            "members": col3.split(","),
        }
    return wgs


def read_module_data(wb):
    from xdrawio.arch.datatypes import Module

    ws = wb["Modules"]
    for tbl in ws._tables:
        if tbl.name == "Modules":
            data = ws[tbl.ref]
            break

    mdls = []
    header = None
    for row in data:
        # Get a list of all columns in each row
        cols = []
        for col in row:
            cols.append(col.value)

        if header is None:
            header = cols
        else:
            mdl = Module(cols[2], cols[4], cols[6])
            mdl.team = cols[0]
            mdl.group = cols[1]
            mdl.wg_type = cols[3]
            mdl.sub_group = cols[5]
            mdls.append(mdl)

    return mdls


def read_data(file_path):
    # Reading an excel file using Python
    import openpyxl

    d = Data()

    # To open Workbook
    wb = openpyxl.load_workbook(file_path)

    d.configurations = xdrawio.xutils.read_configuration_data(wb)

    ws = wb["Teams"]

    # For row 0 and column 0
    for tbl in ws._tables:
        # print(tbl.name)
        # Grab the 'data' from the table
        data = ws[tbl.ref]
        if tbl.name == "Teams":
            d.domains = read_domain_data(data)

        if tbl.name == "WorkGroups":
            d.workgroups = read_workgroup_data(data)

        if tbl.name == "Groups":
            d.groups = read_group_data(data)

        if tbl.name == "LayoutSpecs":
            d.layoutspec = read_layout_specification(data)

    mdls = read_module_data(wb)
    # normalize data
    for mdl in mdls:
        mdl.wg_stype = "WGStyle%d" % (d.workgroups[mdl.wg_type]["type"])

    return mdls, d
