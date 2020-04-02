from xdrawio.utils import randomString

__version_info__ = (0,1,0)
__version__ = '.'.join([str(__value) for __value in __version_info__])
__copyright__ = '2020, NGUYEN Huu Hoa'
__license__ = 'MIT'

def read_column_data(wb):
    ws = wb["Banks"]
    for tbl in ws._tables:
        if tbl.name == "Stages":
            data = ws[tbl.ref]
            break

    td = []
    header = None
    for row in data:
        # Get a list of all columns in each row
        cols = []
        for col in row:
            cols.append(col.value)

        if header is None:
            header = cols
            continue

        td.append({header[i]:cols[i] for i in range(len(cols))})
    
    td.sort(key=lambda x: x["Sort Order"], reverse=False)
    return td


class Data(object):
    def __init__(self):
        super().__init__()
        self.column_stage = {}
        self.banks = {}
        self.styles = {}
        self.configurations = {}


def read_banks_data(wb):
    ws = wb["Banks"]
    for tbl in ws._tables:
        if tbl.name == "Banks":
            data = ws[tbl.ref]
            break

    d = {}
    header = None
    for row in data:
        # Get a list of all columns in each row
        cols = []
        for col in row:
            cols.append(col.value)

        if header is None:
            header = cols
            continue

        bank_name = cols[0]
        info = {header[i]:cols[i] for i in range(1, len(cols))}
        bank = {}
        bank["id"] = randomString()
        bank["display_name"] = bank_name
        bank["info"] = info
        d[bank_name] = bank

    return d


def read_configuration_data(wb):
    ws = wb["Configuration"]
    for tbl in ws._tables:
        if tbl.name == "Configuration":
            data = ws[tbl.ref]
            break

    cfg = {}
    header = None
    for row in data:
        # Get a list of all columns in each row
        cols = []
        for col in row:
            cols.append(col.value)

        if header is None:
            header = cols
        else:
            cfg[cols[0]] = cols[1]

    return cfg


def read_data(file_path):
    # Reading an excel file using Python 
    import openpyxl

    d = Data()

    # To open Workbook 
    wb = openpyxl.load_workbook(file_path)

    d.configurations = read_configuration_data(wb)
    d.column_stage = read_column_data(wb)
    d.banks = read_banks_data(wb)

    return d


def flatten_data(data):
    all_items = []

    start_x = 0
    start_y = 270
    counter = 0
    stages = {}

    item_padding = 70
    item_height = 40
    x = start_x
    all_items.append({
        "type": "frame",
        "w": len(data.column_stage) * item_padding,
    })
    for step in data.column_stage:
        all_items.append({
            "id": "step-%d" % (counter),
            "display_name": step["Display Name"],
            "x": start_x + 22.5 + counter * item_padding,
            "type": "step_label",
            "h": len(data.banks) * item_height + 40,
        })
        if step["Stage"] not in stages:
            stages[step["Stage"]] = {
                "id": "stage-" + step["Stage"],
                "display_name": step["Stage"].upper(), 
                "x": x,
                "w": item_padding,
                "type": "stage",
                "style": data.configurations["Stage_%s" % step["Stage"]],
            }
        else:
            ss = stages[step["Stage"]]
            ss["w"] = ss["w"] + item_padding
        x += item_padding
        counter += 1

    for ss in stages.values():
        all_items.append(ss)

    for b in data.banks.values():
        all_items.append({
            "id": b["id"],
            "y": start_y,
            "type": "bank"
        })
        bank = {
            "x": 0,
            "w": 100,
            "h": 20,
            "type": "label",
            "parentid": b["id"],
            "w": 120 + len(data.column_stage) * item_padding + 40
        }
        bank.update(b)
        all_items.append(bank)
        counter = 0
        info = b["info"]
        all_items.append({
            "type": "bank-line",
            "parentid": b["id"],
            "w": len(data.column_stage) * item_padding + 40
        })
        status = info["Status"]
        for step in data.column_stage:
            if status == "Completed":
                value = ""
            else:
                value = info[step["Step"]]

            if value is None:
                style = data.configurations["StepStatus_NA"]
                value = ""
            else:
                style = data.configurations["StepStatus_%s" % step["Stage"]]

            all_items.append({
                "id": "%s-%d" % (b["id"], counter),
                "display_name": value,
                "x": 175 + counter * item_padding,
                "type": "step",
                "parentid": b["id"],
                "style": style,
            })
            counter += 1

        start_y += item_height

    return all_items
